#!/usr/bin/env python
"""Download the DOL OFLC LCA disclosure file (Akamai-gated to bots, but a real
browser session passes) and convert it to CSV for the radar's DOL importer.

Usage:
  python scout/scout_dol.py                 # latest quarter it can find
  python scout/scout_dol.py --url <xlsx>    # explicit file

Writes radar/data/dol-raw/<name>.csv. Then:
  npm run radar:import-dol -- radar/data/dol-raw/<name>.csv
  npm run radar:enrich
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from radar_scout.logging_utils import configure_logging, get_logger
from radar_scout.net import UA

log = get_logger("scout_dol")

RADAR_PATH = Path(__file__).resolve().parents[1]
PERF_URL = "https://www.dol.gov/agencies/eta/foreign-labor/performance"
LCA_LINK_RE = re.compile(r"LCA_Dis?l?closure_Data_FY\d{4}_Q\d\.xlsx", re.IGNORECASE)


def _fy_q_key(url: str) -> tuple[int, int]:
    m = re.search(r"FY(\d{4})_Q(\d)", url, re.IGNORECASE)
    return (int(m.group(1)), int(m.group(2))) if m else (0, 0)


def find_latest_lca_url(page) -> str | None:
    page.goto(PERF_URL, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(2500)
    candidates = []
    for a in page.query_selector_all("a"):
        href = a.get_attribute("href") or ""
        if LCA_LINK_RE.search(href):
            candidates.append(href if href.startswith("http") else "https://www.dol.gov" + href)
    if not candidates:
        return None
    # newest fiscal quarter, and prefer the /media/ CDN host (it serves to
    # browser sessions; the /sites/ path 403s more aggressively)
    candidates.sort(key=lambda u: (_fy_q_key(u), "/media/" in u), reverse=True)
    return candidates[0]


def xlsx_to_csv(xlsx_path: Path, csv_path: Path) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    rows = 0
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
            rows += 1
            if rows % 100000 == 0:
                log.info("converting", rows=rows)
    wb.close()
    return rows


def main() -> int:
    configure_logging()
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--url", default=None)
    parser.add_argument("--keep-xlsx", action="store_true")
    args = parser.parse_args()

    out_dir = RADAR_PATH / "radar" / "data" / "dol-raw"
    out_dir.mkdir(parents=True, exist_ok=True)

    from playwright.sync_api import sync_playwright

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context(user_agent=UA, accept_downloads=True)
        page = context.new_page()

        url = args.url or find_latest_lca_url(page)
        if not url:
            log.error("no_lca_link_found")
            return 1
        log.info("downloading", url=url)

        name = url.rsplit("/", 1)[-1]
        xlsx_path = out_dir / name
        # Trigger a real browser download so Akamai sees a full browser
        # fingerprint (the request API alone gets a 403)
        try:
            with page.expect_download(timeout=600000) as download_info:
                try:
                    page.goto(url, timeout=600000)
                except Exception:
                    # A navigation that resolves to a file download is aborted
                    # by Chromium (net::ERR_ABORTED); the download event still fires
                    pass
            download = download_info.value
            download.save_as(str(xlsx_path))
        except Exception as error:
            log.error("download_failed", error=str(error)[:120])
            return 1
        log.info("downloaded", path=str(xlsx_path), mb=round(xlsx_path.stat().st_size / 1048576, 1))
        browser.close()

    csv_path = out_dir / (xlsx_path.stem + ".csv")
    rows = xlsx_to_csv(xlsx_path, csv_path)
    log.info("converted", csv=str(csv_path), rows=rows)
    if not args.keep_xlsx:
        xlsx_path.unlink(missing_ok=True)
    print(str(csv_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
